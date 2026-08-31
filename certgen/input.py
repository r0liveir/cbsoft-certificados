from __future__ import annotations

import io
import re
import urllib.parse
import urllib.request
from pathlib import Path

from .core import CertificateError


def _download(url: str) -> bytes:
    parsed = urllib.parse.urlparse(url)
    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if match:
        url = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"
    try:
        with urllib.request.urlopen(url, timeout=30) as response:
            return response.read()
    except OSError as error:
        raise CertificateError(f"Could not download spreadsheet: {error}") from error


def _workbook(source: str):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise CertificateError("Install dependencies first: python -m pip install -e .") from error
    data = _download(source) if source.startswith(("https://", "http://")) else Path(source).read_bytes()
    try:
        return load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as error:
        raise CertificateError("Input must be a valid .xlsx workbook or public Google Sheets URL.") from error


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _rows(sheet) -> list[tuple[str, ...]]:
    return [tuple(_text(cell) for cell in row) for row in sheet.iter_rows(values_only=True)]


def read_xlsx(source: str, config_sheet: str | None, data_sheet: str) -> tuple[dict[str, str], list[dict[str, str]]]:
    workbook = _workbook(source)
    if data_sheet not in workbook.sheetnames:
        raise CertificateError(f"Data sheet {data_sheet!r} not found. Available: {', '.join(workbook.sheetnames)}")

    config: dict[str, str] = {}
    if config_sheet:
        if config_sheet not in workbook.sheetnames:
            raise CertificateError(f"Config sheet {config_sheet!r} not found.")
        for index, values in enumerate(_rows(workbook[config_sheet])[1:], start=2):
            if not any(values):
                continue
            if len(values) < 2 or not values[0] or not values[1]:
                raise CertificateError(f"{config_sheet}!{index} needs a key and value.")
            if values[0] in config:
                raise CertificateError(f"Duplicate config key {values[0]!r}.")
            config[values[0]] = values[1]

    values = _rows(workbook[data_sheet])
    if not values or not any(values[0]):
        raise CertificateError(f"{data_sheet!r} needs a header row.")
    headers = list(values[0])
    if not all(headers) or len(set(headers)) != len(headers):
        raise CertificateError(f"{data_sheet!r} headers must be non-empty and unique.")
    rows = [dict(zip(headers, row)) for row in values[1:] if any(row)]
    if not rows:
        raise CertificateError(f"{data_sheet!r} has no data rows.")
    return config, rows
